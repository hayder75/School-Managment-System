#!/usr/bin/env python3
"""Extract all Mount Olive School xlsx files into normalized JSON for seeding."""
import json
import os
import re
import openpyxl

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'parsed')
os.makedirs(OUT, exist_ok=True)


def clean(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def to_num(v):
    if v is None:
        return None
    s = str(v).strip().replace(',', '')
    if not s:
        return None
    try:
        f = float(s)
        if f == int(f):
            return int(f)
        return round(f, 2)
    except ValueError:
        return None


def main():
    # ── 1. BASIC STUDENT PRIMARY ──
    wb = openpyxl.load_workbook('basic stu prim (2).xlsx', read_only=True, data_only=True)
    ws = wb['sheet1']
    rows = ws.iter_rows(values_only=True)
    header = [clean(h) for h in next(rows)]
    students = []
    for r in rows:
        rec = {header[i]: r[i] for i in range(min(len(header), len(r)))}
        if not rec.get('Student ID'):
            continue
        students.append({
            'student_id': clean(rec.get('Student ID')),
            'institution_id': clean(rec.get('Institution ID')),
            'first_name': clean(rec.get('First Name')),
            'father_name': clean(rec.get("Father's Name")),
            'grandfather_name': clean(rec.get("Grandfather's Name")),
            'admission_type': clean(rec.get('Admission Type')),
            'sex': clean(rec.get('Sex')),
            'disability': str(clean(rec.get('Disability'))).lower() in ('yes', 'true', '1'),
            'disability_type': clean(rec.get('Disability Type')),
            'dob': clean(rec.get('Date Of Birth')),
            'nationality': clean(rec.get('Nationality')),
            'family_kebele': clean(rec.get('Family Kebele')),
            'location_type': clean(rec.get('Location Type')),
            'father_education': clean(rec.get("Father's Or Male Guardian's Education Level*")),
            'mother_education': clean(rec.get("Mother's Or Female Guardian's Education Level*")),
            'economic_status': clean(rec.get('Student Economic Status')),
            'guardian_name': clean(rec.get("Parent's or Guardian's Full Name")),
            'family_head_gender': clean(rec.get("Family Head's Gender")),
            'guardian_email': clean(rec.get("Parent's or Guardian's Email")),
            'guardian_phone': clean(rec.get("Parent's or Guardian's Phone")),
            'national_id': clean(rec.get('National ID')),
            'region_of_residence': clean(rec.get('Region of Residence')),
            'zone_of_residence': clean(rec.get('Zone of Residence')),
            'woreda_of_residence': clean(rec.get('Woreda of Residence')),
            'region_of_birth': clean(rec.get('Region of Birth')),
            'zone_of_birth': clean(rec.get('Zone of Birth')),
            'woreda_of_birth': clean(rec.get('Woreda of Birth')),
            'parent_status': clean(rec.get('Parent Status')),
            'country_of_birth': clean(rec.get('Country of Birth')),
        })
    wb.close()
    with open(os.path.join(OUT, 'students_primary.json'), 'w') as f:
        json.dump(students, f, indent=1, ensure_ascii=False)
    print(f'primary students: {len(students)}')

    # ── 2. ENROLLMENT PRIMARY ──
    wb = openpyxl.load_workbook('enroll prim (2).xlsx', read_only=True, data_only=True)
    ws = wb['sheet1']
    rows = ws.iter_rows(values_only=True)
    header = [clean(h) for h in next(rows)]
    enroll_prim = []
    for r in rows:
        rec = {header[i]: r[i] for i in range(min(len(header), len(r)))}
        if not rec.get('Student ID'):
            continue
        enroll_prim.append({
            'student_id': clean(rec.get('Student ID')),
            'academic_year': clean(rec.get('Academic Year')),
            'admission_category': clean(rec.get('Admission Category')),
            'admission_modality': clean(rec.get('Admission Modality')),
            'grade_level': clean(rec.get('Grade Level')),
            'section': clean(rec.get('Section')),
            'education_stream': clean(rec.get('Education Stream')),
            'cte_field_1': clean(rec.get('Career and Technical Education 1st Field')),
            'cte_field_2': clean(rec.get('Career and Technical Education 2st Field')),
            'num_textbooks': to_num(rec.get('Nbr of Text Books')),
            'instructional_language': clean(rec.get('Main Instructional Language')),
            'school_feeding': str(clean(rec.get('Participation in School Feeding'))).lower() in ('yes', 'true', '1'),
            'food_ration_home': str(clean(rec.get('Food Ration Home Taking*'))).lower() in ('yes', 'true', '1'),
            'meals_per_week': to_num(rec.get('Nbr of Meals per Week')),
        })
    wb.close()
    with open(os.path.join(OUT, 'enroll_primary.json'), 'w') as f:
        json.dump(enroll_prim, f, indent=1, ensure_ascii=False)
    print(f'primary enrollments: {len(enroll_prim)}')

    # ── 3. ENROLLMENT KG ──
    wb = openpyxl.load_workbook('enroll kg (2).xlsx', read_only=True, data_only=True)
    ws = wb['sheet1']
    rows = ws.iter_rows(values_only=True)
    header = [clean(h) for h in next(rows)]
    enroll_kg = []
    for r in rows:
        rec = {header[i]: r[i] for i in range(min(len(header), len(r)))}
        if not rec.get('Student ID'):
            continue
        enroll_kg.append({
            'student_id': clean(rec.get('Student ID')),
            'institution_id': clean(rec.get('Institution ID')),
            'first_name': clean(rec.get('First Name')),
            'academic_year': clean(rec.get('Academic Year')),
            'admission_category': clean(rec.get('Admission Category')),
            'admission_modality': clean(rec.get('Admission Modality')),
            'grade_level': clean(rec.get('Grade Level')),
            'section': clean(rec.get('Section')),
            'education_stream': clean(rec.get('Education Stream')),
            'cte_field_1': clean(rec.get('Career and Technical Education 1st Field')),
            'cte_field_2': clean(rec.get('Career and Technical Education 2st Field')),
            'num_textbooks': to_num(rec.get('Nbr of Text Books')),
            'instructional_language': clean(rec.get('Main Instructional Language')),
            'school_feeding': str(clean(rec.get('Participation in School Feeding'))).lower() in ('yes', 'true', '1'),
            'food_ration_home': str(clean(rec.get('Food Ration Home Taking*'))).lower() in ('yes', 'true', '1'),
            'meals_per_week': to_num(rec.get('Nbr of Meals per Week')),
        })
    wb.close()
    with open(os.path.join(OUT, 'enroll_kg.json'), 'w') as f:
        json.dump(enroll_kg, f, indent=1, ensure_ascii=False)
    print(f'KG enrollments: {len(enroll_kg)}')

    # ── 4. STAFF (statistical data) ──
    wb = openpyxl.load_workbook('statistical data (2).xlsx', read_only=True, data_only=True)
    staff = {}

    def sheet_rows(name, cols):
        ws = wb[name]
        out = []
        for r in ws.iter_rows(values_only=True):
            rec = [clean(x) for x in r]
            out.append(rec)
        return out

    # primary teachers: rows start at r4 (idx3), columns: no,name,sex,qual,field,subject,class,sec_count,reg,ot,total
    prim_teachers = []
    for r in sheet_rows('primary teachers', 11)[3:]:
        if not r[0] or not str(r[0]).isdigit():
            continue
        prim_teachers.append({
            'no': r[0], 'name': r[1], 'sex': r[2], 'qualification': r[3],
            'field_of_study': r[4], 'subject': r[5], 'classes': r[6],
            'section_count': to_num(r[7]), 'reg_periods': to_num(r[8]),
            'overtime_periods': to_num(r[9]), 'total_periods': to_num(r[10]),
        })
    staff['primary_teachers'] = prim_teachers
    print(f'primary teachers: {len(prim_teachers)}')

    # kG staff: rows start at r4 (idx3), cols: no,name,sex,qual,field,subject,class,sec_count,periods,remark
    kg_staff = []
    for r in sheet_rows('kG staff', 10)[3:]:
        if not r[0] or not str(r[0]).isdigit():
            continue
        kg_staff.append({
            'no': r[0], 'name': r[1], 'sex': r[2], 'qualification': r[3],
            'field_of_study': r[4], 'subject': r[5], 'classes': r[6],
            'section_count': to_num(r[7]), 'periods': to_num(r[8]), 'remark': r[9],
        })
    staff['kg_staff'] = kg_staff
    print(f'KG staff: {len(kg_staff)}')

    # supportive: rows start at r4 (idx3), cols: no,name,sex,qual,field,position,remark
    supportive = []
    for r in sheet_rows('supportive', 7)[3:]:
        if not r[0] or not str(r[0]).isdigit():
            continue
        supportive.append({
            'no': r[0], 'name': r[1], 'sex': r[2], 'qualification': r[3],
            'field_of_study': r[4], 'position': r[5], 'remark': r[6],
        })
    staff['supportive'] = supportive
    print(f'supportive: {len(supportive)}')

    # management: rows start at r4 (idx3), cols: no,name,sex,qual,subject,position,remark
    management = []
    for r in sheet_rows('management', 7)[3:]:
        if not r[0] or not str(r[0]).isdigit():
            continue
        management.append({
            'no': r[0], 'name': r[1], 'sex': r[2], 'qualification': r[3],
            'subject': r[4], 'position': r[5], 'remark': r[6],
        })
    staff['management'] = management
    print(f'management: {len(management)}')

    with open(os.path.join(OUT, 'staff.json'), 'w') as f:
        json.dump(staff, f, indent=1, ensure_ascii=False)
    wb.close()

    # ── 5. PAYROLL (June 2018) ──
    wb = openpyxl.load_workbook('JUNE SALARY 2018.xlsx', read_only=True, data_only=True)

    def payroll_rows(sheet):
        ws = wb[sheet]
        out = []
        for r in ws.iter_rows(values_only=True):
            if r[0] is not None and str(r[0]).strip().isdigit() and r[1] and str(r[1]).strip():
                out.append([clean(x) for x in r])
        return out

    def asnum(r, i):
        return to_num(r[i]) if i < len(r) else None

    def sval(r, i):
        return clean(r[i]) if i < len(r) else None

    payroll = {'Teachers': [], 'KG_Supporting': [], 'SUPPORT': [], 'KG_Teachers': [], 'ADMIN_1': [], 'Admin_2': []}

    # Teachers sheet: c1 name, c2 job, c3 basic, c4 workdays, c5 abday, c7 basic2, c8 transp, c9 ot, c10 backpay, c11 unitleader, c12 dh, c13 gross, c14 taxable, c15 incometax, c16 schoolpay, c17 eder, c18 office loan, c19 cafe loan, c20 pen7, c21 pen11, c22 nestarving, c23 totalded, c24 netpay
    for r in payroll_rows('Teachers'):
        payroll['Teachers'].append({
            'name': r[1], 'job_title': sval(r, 2), 'basic_salary': asnum(r, 3),
            'work_days': asnum(r, 4), 'absent_days': asnum(r, 5),
            'transport_allowance': asnum(r, 8), 'overtime': asnum(r, 9), 'back_pay': asnum(r, 10),
            'unit_leader_allowance': asnum(r, 11), 'department_head_allowance': asnum(r, 12),
            'gross': asnum(r, 13), 'taxable': asnum(r, 14), 'income_tax': asnum(r, 15),
            'school_pay': asnum(r, 16), 'eder': asnum(r, 17), 'office_loan': asnum(r, 18),
            'cafe_loan': asnum(r, 19), 'pension_employee': asnum(r, 20), 'pension_employer': asnum(r, 21),
            'ne_starving': asnum(r, 22), 'total_deductions': asnum(r, 23), 'net_pay': asnum(r, 24),
        })

    # KG,Supporting: c1 name, c2 job, c3 basic, c4 workdays, c7 basic2, c8 backpay, c9 transp, c12 incometax, c13 office loan, c14 schoolpay, c15 pen7, c16 pen11, c17 nestarving, c18 totalded, c19 netpay
    for r in payroll_rows('KG,Supporting'):
        payroll['KG_Supporting'].append({
            'name': r[1], 'job_title': sval(r, 2), 'basic_salary': asnum(r, 3),
            'work_days': asnum(r, 4),
            'back_pay': asnum(r, 8), 'transport_allowance': asnum(r, 9),
            'income_tax': asnum(r, 12), 'office_loan': asnum(r, 13), 'school_pay': asnum(r, 14),
            'pension_employee': asnum(r, 15), 'pension_employer': asnum(r, 16), 'ne_starving': asnum(r, 17),
            'total_deductions': asnum(r, 18), 'net_pay': asnum(r, 19),
        })

    # SUPPORT: c1 name, c2 job, c3 basic, c4 workdays, c5 basic2, c8 transp, c11 incometax, c12 eder, c13 cafe loan, c14 schoolpay, c15 office loan, c16 pen7, c17 pen11, c18 nestarving, c19 totalded, c20 netpay
    for r in payroll_rows('SUPPORT'):
        payroll['SUPPORT'].append({
            'name': r[1], 'job_title': sval(r, 2), 'basic_salary': asnum(r, 3),
            'work_days': asnum(r, 4),
            'transport_allowance': asnum(r, 8), 'income_tax': asnum(r, 11), 'eder': asnum(r, 12),
            'cafe_loan': asnum(r, 13), 'school_pay': asnum(r, 14), 'office_loan': asnum(r, 15),
            'pension_employee': asnum(r, 16), 'pension_employer': asnum(r, 17), 'ne_starving': asnum(r, 18),
            'total_deductions': asnum(r, 19), 'net_pay': asnum(r, 20),
        })

    # KG TEACHERS: c1 name, c2 job, c3 basic, c4 workdays, c5 basic2, c6 backpay, c9 incometax, c10 schoolpay, c11 office loan, c12 pen7, c13 pen11, c14 nestarving, c15 totalded, c16 netpay
    for r in payroll_rows('KG TEACHERS'):
        payroll['KG_Teachers'].append({
            'name': r[1], 'job_title': sval(r, 2), 'basic_salary': asnum(r, 3),
            'work_days': asnum(r, 4),
            'back_pay': asnum(r, 6), 'income_tax': asnum(r, 9), 'school_pay': asnum(r, 10),
            'office_loan': asnum(r, 11), 'pension_employee': asnum(r, 12), 'pension_employer': asnum(r, 13),
            'ne_starving': asnum(r, 14), 'total_deductions': asnum(r, 15), 'net_pay': asnum(r, 16),
        })

    # ADMIN 1: c1 name, c2 job, c3 workdays, c4 basic, c5 transp, c6 hallow, c7 accallow, c10 incometax, c11 pen7, c12 pen11, c13 nestarving, c14 totalded, c15 netpay
    for r in payroll_rows('ADMIN 1'):
        payroll['ADMIN_1'].append({
            'name': r[1], 'job_title': sval(r, 2), 'basic_salary': asnum(r, 4),
            'work_days': asnum(r, 3), 'transport_allowance': asnum(r, 5), 'housing_allowance': asnum(r, 6),
            'account_allowance': asnum(r, 7), 'income_tax': asnum(r, 10), 'pension_employee': asnum(r, 11),
            'pension_employer': asnum(r, 12), 'ne_starving': asnum(r, 13), 'total_deductions': asnum(r, 14),
            'net_pay': asnum(r, 15),
        })

    # Admin 2: c1 name, c2 job, c3 basic, c4 workdays, c6 transp, c7 hallow, c8 accallow, c9 phoallow, c10 ot, c13 incometax, c14 eder, c15 loan, c16 schoolpay, c17 pen7, c18 pen11, c19 totalded, c20 netpay
    for r in payroll_rows('Admin 2'):
        payroll['Admin_2'].append({
            'name': r[1], 'job_title': sval(r, 2), 'basic_salary': asnum(r, 3),
            'work_days': asnum(r, 4), 'transport_allowance': asnum(r, 6), 'housing_allowance': asnum(r, 7),
            'account_allowance': asnum(r, 8), 'phone_allowance': asnum(r, 9), 'overtime': asnum(r, 10),
            'income_tax': asnum(r, 13), 'eder': asnum(r, 14), 'office_loan': asnum(r, 15), 'school_pay': asnum(r, 16),
            'pension_employee': asnum(r, 17), 'pension_employer': asnum(r, 18), 'total_deductions': asnum(r, 19),
            'net_pay': asnum(r, 20),
        })

    for k, v in payroll.items():
        print(f'payroll {k}: {len(v)} rows')
    with open(os.path.join(OUT, 'payroll.json'), 'w') as f:
        json.dump(payroll, f, indent=1, ensure_ascii=False)
    wb.close()

    print(f'\nAll data extracted to {OUT}')


if __name__ == '__main__':
    main()
